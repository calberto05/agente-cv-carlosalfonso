"""
Punto de entrada del framework de evaluación.

Uso:
    python -m evaluation.main                          # todos los casos
    python -m evaluation.main --category precision     # filtrar por categoría
    python -m evaluation.main --test hackathones_mencionados  # un caso
    python -m evaluation.main --excel                  # además, exporta a Excel

Requiere:
    - AGENT_ENDPOINT (o el default de Cloud Run)
    - AGENT_API_KEY
    - GOOGLE_CLOUD_PROJECT + GOOGLE_GENAI_USE_VERTEXAI=true (o GOOGLE_API_KEY)
"""

import argparse
import os
import sys
import time
from dataclasses import dataclass
from datetime import datetime

# Windows (cp1252) no soporta los caracteres Unicode que usa rich (→, ✓, ✗,
# bordes de tabla). Sin esto, la consola revienta con UnicodeEncodeError a
# mitad de la corrida.
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")

from rich.console import Console
from rich.table import Table
from rich import box
from rich.panel import Panel
from rich.text import Text
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from evaluation.test_cases import TEST_CASES, TestCase
from evaluation.runner import run_test_case
from evaluation.judge import evaluate

console = Console()

PASS_THRESHOLD = 3  # overall_score mínimo para considerar el test como pasado


@dataclass
class TestResult:
    test: TestCase
    responses: list[str]
    evaluation: dict
    error: str | None = None

    @property
    def passed(self) -> bool:
        if self.error:
            return False
        return self.evaluation.get("overall_score", 0) >= PASS_THRESHOLD

    @property
    def score(self) -> int:
        return self.evaluation.get("overall_score", 0) if not self.error else 0


def run_single(test: TestCase) -> TestResult:
    console.print(f"  [dim]Ejecutando:[/dim] [bold]{test.name}[/bold]", end=" ")
    try:
        responses = run_test_case(test)
        evaluation = evaluate(
            messages=test.messages,
            responses=responses,
            criteria=test.criteria,
            expected_behavior=test.expected_behavior,
        )
        result = TestResult(test=test, responses=responses, evaluation=evaluation)
        status = "[green]PASS[/green]" if result.passed else "[red]FAIL[/red]"
        console.print(f"→ {status} (score {result.score}/5)")
        return result
    except Exception as exc:
        console.print(f"→ [red]ERROR[/red] ({exc})")
        return TestResult(test=test, responses=[], evaluation={}, error=str(exc))


def print_detail(result: TestResult) -> None:
    color = "green" if result.passed else "red"
    title = f"[{color}]{result.test.name}[/{color}]  (categoría: {result.test.category})"
    console.print(Panel(title, expand=False))

    if result.error:
        console.print(f"  [red]Error:[/red] {result.error}")
        return

    # Conversación
    for i, (msg, resp) in enumerate(zip(result.test.messages, result.responses), 1):
        console.print(f"  [cyan]T{i} Usuario:[/cyan] {msg}")
        console.print(f"  [cyan]T{i} Agente:[/cyan]  {resp[:300]}{'...' if len(resp) > 300 else ''}")
        console.print()

    # Criterios
    for cr in result.evaluation.get("criteria_results", []):
        icon = "[green]✓[/green]" if cr["passed"] else "[red]✗[/red]"
        console.print(f"  {icon} {cr['criterion']}")
        console.print(f"     [dim]{cr['reasoning']}[/dim]")

    console.print(f"\n  [bold]Resumen:[/bold] {result.evaluation.get('summary', '')}")
    console.print(f"  [bold]Score:[/bold] {result.score}/5\n")


def print_summary(results: list[TestResult]) -> None:
    table = Table(
        title="Resumen de evaluación",
        box=box.ROUNDED,
        show_lines=True,
    )
    table.add_column("Test", style="bold", no_wrap=True)
    table.add_column("Categoría")
    table.add_column("Score", justify="center")
    table.add_column("Estado", justify="center")

    for r in results:
        score_str = f"{r.score}/5" if not r.error else "ERR"
        status = "[green]PASS[/green]" if r.passed else "[red]FAIL[/red]"
        table.add_row(r.test.name, r.test.category, score_str, status)

    console.print(table)

    total = len(results)
    passed = sum(1 for r in results if r.passed)
    errors = sum(1 for r in results if r.error)
    avg_score = sum(r.score for r in results if not r.error) / max(total - errors, 1)

    console.print()
    console.print(f"  Tests ejecutados : {total}")
    console.print(f"  Pasados          : [green]{passed}[/green]")
    console.print(f"  Fallidos         : [red]{total - passed - errors}[/red]")
    if errors:
        console.print(f"  Errores          : [yellow]{errors}[/yellow]")
    console.print(f"  Score promedio   : {avg_score:.1f}/5")
    console.print()


def export_to_excel(results: list[TestResult], path: str) -> None:
    """Exporta los resultados a un .xlsx, una fila por criterio evaluado."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    headers = [
        "Test", "Categoría", "Pregunta", "Respuesta", "Criterio",
        "Aprobado", "Razonamiento", "Score", "Estado",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    for col, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center")

    pass_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    fail_fill = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    for r in results:
        pregunta = "\n".join(f"T{i}: {m}" for i, m in enumerate(r.test.messages, 1))
        respuesta = "\n".join(f"T{i}: {resp}" for i, resp in enumerate(r.responses, 1))
        estado = "ERROR" if r.error else ("PASS" if r.passed else "FAIL")

        if r.error:
            ws.append([
                r.test.name, r.test.category, pregunta, respuesta,
                "", "", r.error, "", estado,
            ])
            ws.cell(row=ws.max_row, column=9).fill = fail_fill
            continue

        criteria_results = r.evaluation.get("criteria_results", [])
        if not criteria_results:
            ws.append([r.test.name, r.test.category, pregunta, respuesta, "", "", "", r.score, estado])
            continue

        for cr in criteria_results:
            aprobado = "Sí" if cr.get("passed") else "No"
            ws.append([
                r.test.name, r.test.category, pregunta, respuesta,
                cr.get("criterion", ""), aprobado, cr.get("reasoning", ""),
                r.score, estado,
            ])
            ws.cell(row=ws.max_row, column=6).fill = pass_fill if cr.get("passed") else fail_fill

    for col_idx, width in zip(range(1, 10), [22, 14, 40, 45, 40, 10, 45, 8, 10]):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "A2"

    wb.save(path)
    console.print(f"  [bold green]Excel generado:[/bold green] {path}\n")


def main() -> None:
    parser = argparse.ArgumentParser(description="Evaluación del agente CV")
    parser.add_argument("--category", help="Filtrar por categoría")
    parser.add_argument("--test", help="Ejecutar un caso por nombre")
    parser.add_argument("--detail", action="store_true", help="Mostrar detalle completo")
    parser.add_argument(
        "--excel",
        nargs="?",
        const="__auto__",
        default=None,
        help="Exporta los resultados a un .xlsx (opcionalmente indica la ruta)",
    )
    args = parser.parse_args()

    cases = TEST_CASES
    if args.category:
        cases = [t for t in cases if t.category == args.category]
    if args.test:
        cases = [t for t in cases if t.name == args.test]

    if not cases:
        console.print("[red]No se encontraron casos de prueba con los filtros dados.[/red]")
        sys.exit(1)

    console.print(Panel(
        f"[bold]Evaluación del agente CV[/bold]\n{len(cases)} casos de prueba",
        expand=False,
    ))
    console.print()

    results: list[TestResult] = []
    t0 = time.time()

    for test in cases:
        result = run_single(test)
        results.append(result)

    elapsed = time.time() - t0
    console.print(f"\n  Tiempo total: {elapsed:.1f}s\n")

    if args.detail:
        console.rule("Detalle por caso")
        for r in results:
            print_detail(r)

    console.rule("Resumen")
    print_summary(results)

    if args.excel is not None:
        path = args.excel
        if path == "__auto__":
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            path = f"evaluation/reports/eval_{timestamp}.xlsx"
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        export_to_excel(results, path)

    # Exit code no-zero si algún test falló
    if any(not r.passed for r in results):
        sys.exit(1)


if __name__ == "__main__":
    main()
